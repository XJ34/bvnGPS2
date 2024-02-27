from itertools import chain

import numpy as np
import openpyxl
import pandas as pd
import xlrd


def load_pathway():
    pathway = "c2.cp.kegg.v7.2.symbols.gmt"
    gene_in_pathways = []
    for line in open(pathway):
        line = line.strip()
        split_result = line.split('\t')
        gene_in_pathway = split_result[2:]
        gene_in_pathways.append(gene_in_pathway)
    return gene_in_pathways


def get_col_value(wsheet, col):
    rows = wsheet.max_row
    col_data = []
    for i in range(2, rows + 1):
        cell_value = wsheet.cell(i, col).value
        col_data.append(cell_value)
    return col_data


def comb_mine(n, k):
    up = np.math.factorial(n)
    down = np.math.factorial(k) * np.math.factorial((n - k))
    return up // down


label6269_1 = list(chain.from_iterable(pd.read_excel('label6269-1.xlsx', usecols=[2]).values))
print(label6269_1)
print(len(label6269_1))
label6269_2 = list(chain.from_iterable(pd.read_excel('label6269-2.xlsx', usecols=[2]).values))
print(len(label6269_2))
label6269_3 = list(chain.from_iterable(pd.read_excel('label6269-3.xlsx', usecols=[2]).values))
print(len(label6269_3))
label11755 = list(chain.from_iterable(pd.read_excel('label11755.xlsx', usecols=[2]).values))
print(len(label11755))
label13015 = list(chain.from_iterable(pd.read_excel('label13015.xlsx', usecols=[2]).values))
label13015 = list(filter(lambda x: x != 10, label13015))
print(len(label13015))
label13015_2 = list(chain.from_iterable(pd.read_excel('label13015-2.xlsx', usecols=[2]).values))
label13015_2 = list(filter(lambda x: x != 10, label13015_2))
print(len(label13015_2))
label16129 = list(chain.from_iterable(pd.read_excel('label16129.xlsx', usecols=[2]).values))
print(len(label16129))
label16129_2 = list(chain.from_iterable(pd.read_excel('label16129-2.xlsx', usecols=[2]).values))
print(len(label16129_2))
label16129_3 = list(chain.from_iterable(pd.read_excel('label16129-3.xlsx', usecols=[2]).values))
print(len(label16129_3))
label17156 = list(chain.from_iterable(pd.read_excel('label17156.xlsx', usecols=[2]).values))
print(len(label17156))
label22098 = list(chain.from_iterable(pd.read_excel('label22098.xlsx', usecols=[2]).values))
label22098 = list(filter(lambda x: x != 10, label22098))
print(len(label22098))
label23140 = list(chain.from_iterable(pd.read_excel('label23140.xlsx', usecols=[2]).values))
print(len(label23140))
label25504 = list(chain.from_iterable(pd.read_excel('label25504.xlsx', usecols=[2]).values))
label25504 = list(filter(lambda x: x != 14, label25504))
print(len(label25504))
label25504_3 = list(chain.from_iterable(pd.read_excel('label25504-3.xlsx', usecols=[2]).values))
label25504_3 = list(filter(lambda x: x != 3, label25504_3))
print(len(label25504_3))
label25504_4 = list(chain.from_iterable(pd.read_excel('label25504-4.xlsx', usecols=[2]).values))
label25504_4 = list(filter(lambda x: x != 14, label25504_4))
print(len(label25504_4))
label29161 = list(chain.from_iterable(pd.read_excel('label29161.xlsx', usecols=[2]).values))
print(len(label29161))
label33341_2 = list(chain.from_iterable(pd.read_excel('label33341-2.xlsx', usecols=[2]).values))
print(len(label33341_2))
label34205 = list(chain.from_iterable(pd.read_excel('label34205.xlsx', usecols=[2]).values))
print(len(label34205))
label38246 = list(chain.from_iterable(pd.read_excel('label38246.xlsx', usecols=[2]).values))
print(len(label38246))
label38900 = list(chain.from_iterable(pd.read_excel('label38900.xlsx', usecols=[2]).values))
print(len(label38900))
label40586 = list(chain.from_iterable(pd.read_excel('label40586.xlsx', usecols=[2]).values))
print(len(label40586))
label51808 = list(chain.from_iterable(pd.read_excel('label51808.xlsx', usecols=[2]).values))
label51808 = list(filter(lambda x: x != 10, label51808))
print(len(label51808))
label69606 = list(chain.from_iterable(pd.read_excel('label69606.xlsx', usecols=[2]).values))
print(len(label69606))
label42834 = list(chain.from_iterable(pd.read_excel('label42834.xlsx', usecols=[2]).values))
label42834 = list(filter(lambda x: x != 10, label42834))
print(len(label42834))
excel40396 = openpyxl.load_workbook('label40396.xlsx')
sheets1 = excel40396.get_sheet_names()
ws1 = excel40396.get_sheet_by_name(sheets1[0])
label40396 = get_col_value(ws1, 3)

all_need_genes = load_pathway()
all_genes = pd.read_excel('all_data1.xlsx')
print(all_genes.index)
print(all_genes.columns)

label_excel = xlrd.open_workbook('all_labels.xls')
sheet0 = label_excel.sheets()[12]
label68310 = sheet0.col_values(2, 1, )
sheet1 = label_excel.sheets()[11]
label20346 = sheet1.col_values(2, 1, )
print(len(label20346))
sheet2 = label_excel.sheets()[10]
label21802 = sheet2.col_values(2, 1, )
print(len(label21802))
sheet3 = label_excel.sheets()[9]
label27131 = sheet3.col_values(2, 1, )
print(len(label27131))
sheet4 = label_excel.sheets()[8]
label28750 = sheet4.col_values(2, 1, )
print(len(label28750))
sheet5 = label_excel.sheets()[7]
label40012 = sheet5.col_values(2, 1, )
label40012 = list(filter(lambda x: x != 3, label40012))
print(len(label40012))
sheet6 = label_excel.sheets()[6]
label42026 = sheet6.col_values(2, 1, )
print(len(label42026))
sheet7 = label_excel.sheets()[5]
label57065 = sheet7.col_values(2, 1, )
print(len(label57065))
sheet8 = label_excel.sheets()[4]
label60244 = sheet8.col_values(2, 1, )
# label60244 = list(filter(lambda x: x != 3, label60244))
print(len(label60244))
sheet9 = label_excel.sheets()[3]
label63990 = sheet9.col_values(2, 1, )
print(len(label63990))
sheet10 = label_excel.sheets()[2]
label66099 = sheet10.col_values(2, 1, )
print(len(label66099))
sheet11 = label_excel.sheets()[1]
label69528 = sheet11.col_values(2, 1, )
print(len(label69528))
sheet12 = label_excel.sheets()[0]
label111368 = sheet12.col_values(2, 1, )
print(len(label111368))
label9692 = list(chain.from_iterable(pd.read_excel('label9692.xlsx', usecols=[2]).values))
label61821 = list(chain.from_iterable(pd.read_excel('label61821.xlsx', usecols=[2]).values))
label103842 = list(chain.from_iterable(pd.read_excel('label103842.xlsx', usecols=[2]).values))

labels = np.array(
    label6269_1 + label6269_2 + label6269_3 + label11755 + label13015_2 + label16129 + label16129_2 +
    label17156 + label20346 + label22098 + label23140 + label25504 + label25504_3 + label25504_4 + label27131 + label28750 + label29161 + label33341_2
    + label34205 + label38246 + label38900 + label40012 + label40396 + label42026 + label42834 + label51808 + label57065 +
    label60244 + label63990 + label66099 +
    label68310 + label69528 + label69606 + label111368)
print(len(labels))
print(all_genes.index)
label_health = []
label_num = 0
for i in all_genes.index:
    label_health.append(labels[i])
label_health = np.array(label_health)
label_h = []
for i in label_health:
    if i == 0 or i == 4:
        j = 1
    else:
        j = 0
    label_h.append(j)
label_h = np.array(label_h)

label_bac = []
bac_index = []
bac_num = 0
for i in labels:
    if i == 0 or i == 11 or i == 12 or i == 4 or i == 2:
        if bac_num in all_genes.index:
            bac_index.append(bac_num)
            j = i
            label_bac.append(j)
    bac_num = bac_num + 1

label_b = []
for i in label_bac:
    if i == 11:
        j = 1
    else:
        j = 0
    label_b.append(j)
label_b = np.array(label_b)

label_bacn = []
bacn_index = []
bacn_num = 0
for i in labels:
    if i == 0 or i == 11 or i == 12 or i == 4 or i == 2:
        if bacn_num in all_genes.index:
            bacn_index.append(bacn_num)
            j = i
            label_bacn.append(j)
    bacn_num = bacn_num + 1

label_bn = []
for i in label_bacn:
    if i == 12:
        j = 1
    else:
        j = 0
    label_bn.append(j)
label_bn = np.array(label_bn)

label_virus = []
for i in all_genes.index:
    label_virus.append(labels[i])
label_virus = np.array(label_virus)
label_v = []
for i in label_virus:
    if i == 2:
        j = 1
    else:
        j = 0
    label_v.append(j)
label_v = np.array(label_v)

for gene_piece in all_need_genes:
    for gene1 in gene_piece:
        gene_index = gene_piece.index(gene1)
        for j in range(gene_index + 1, len(gene_piece)):
            gene2 = gene_piece[j]
            if gene1 in all_genes.columns and gene2 in all_genes.columns:
                gene1_data = np.array(all_genes[gene1])
                gene2_data = np.array(all_genes[gene2])
                pair = gene1_data > gene2_data
                gene1_positive_data = np.array(all_genes.loc[bac_index, [gene1]].values)
                gene2_positive_data = np.array(all_genes.loc[bac_index, [gene2]].values)
                positive_pair = gene1_positive_data > gene2_positive_data
                gene1_negative_data = np.array(all_genes.loc[bacn_index, [gene1]].values)
                gene2_negative_data = np.array(all_genes.loc[bacn_index, [gene2]].values)
                negative_pair = gene1_negative_data > gene2_negative_data
                positive_pair = positive_pair.flatten()
                negative_pair = negative_pair.flatten()
                a1 = np.sum((pair > 0.5) * (label_h < 0.5))
                b1 = np.sum((pair > 0.5) * (label_h > 0.5))
                c1 = np.sum((pair < 0.5) * (label_h < 0.5))
                d1 = np.sum((pair < 0.5) * (label_h > 0.5))
                n1 = a1 + b1 + c1 + d1
                p1 = comb_mine(a1 + b1, a1) * comb_mine(c1 + d1, c1) / comb_mine(n1, a1 + c1)

                a2 = np.sum((positive_pair > 0.5) * (label_b < 0.5))
                b2 = np.sum((positive_pair > 0.5) * (label_b > 0.5))
                c2 = np.sum((positive_pair < 0.5) * (label_b < 0.5))
                d2 = np.sum((positive_pair < 0.5) * (label_b > 0.5))
                n2 = a2 + b2 + c2 + d2
                p2 = comb_mine(a2 + b2, a2) * comb_mine(c2 + d2, c2) / comb_mine(n2, a2 + c2)

                a3 = np.sum((negative_pair > 0.5) * (label_bn < 0.5))
                b3 = np.sum((negative_pair > 0.5) * (label_bn > 0.5))
                c3 = np.sum((negative_pair < 0.5) * (label_bn < 0.5))
                d3 = np.sum((negative_pair < 0.5) * (label_bn > 0.5))
                n3 = a3 + b3 + c3 + d3
                p3 = comb_mine(a3 + b3, a3) * comb_mine(c3 + d3, c3) / comb_mine(n3, a3 + c3)

                a4 = np.sum((pair > 0.5) * (label_v < 0.5))
                b4 = np.sum((pair > 0.5) * (label_v > 0.5))
                c4 = np.sum((pair < 0.5) * (label_v < 0.5))
                d4 = np.sum((pair < 0.5) * (label_v > 0.5))
                n4 = a4 + b4 + c4 + d4
                p4 = comb_mine(a4 + b4, a4) * comb_mine(c4 + d4, c4) / comb_mine(n4, a4 + c4)

                if p1 < 1e-16:
                    with open('result1.txt', 'a', encoding='utf-8') as r1:
                        r1.write(str(gene1))
                        r1.write('  ')
                        r1.write(str(gene2))
                        r1.write('  ')
                        r1.write(str(p1))
                        r1.write('\n')
                if p2 < 1e-16:
                    with open('result2.txt', 'a', encoding='utf-8') as r1:
                        r1.write(str(gene1))
                        r1.write('  ')
                        r1.write(str(gene2))
                        r1.write('  ')
                        r1.write(str(p1))
                        r1.write('\n')
                if p3 < 1e-16:
                    with open('result3.txt', 'a', encoding='utf-8') as r1:
                        r1.write(str(gene1))
                        r1.write('  ')
                        r1.write(str(gene2))
                        r1.write('  ')
                        r1.write(str(p1))
                        r1.write('\n')
                if p4 < 1e-16:
                    with open('result4.txt', 'a', encoding='utf-8') as r1:
                        r1.write(str(gene1))
                        r1.write('  ')
                        r1.write(str(gene2))
                        r1.write('  ')
                        r1.write(str(p1))
                        r1.write('\n')
