from itertools import chain

import numpy as np
import openpyxl
import pandas as pd
import time
import xlrd
import tensorflow as tf
from sklearn.linear_model import Lasso
from sklearn.manifold import TSNE
seed = int(time.time() * 100) % 399
print(seed)
np.random.seed(seed)
all_health_genes = pd.read_excel('health data.xlsx', header=0)
all_health_genes = all_health_genes.values
print(all_health_genes)
all_positive_genes = pd.read_excel('positive data.xlsx', header=0)
all_positive_genes = all_positive_genes.values
all_negative_genes = pd.read_excel('negative data.xlsx', header=0)  # gram+ 和 gram- 变量名没改
all_negative_genes = all_negative_genes.values
all_virus_genes = pd.read_excel('virus data.xlsx', header=0)
all_virus_genes = all_virus_genes.values


def load_pathway():
    pathway = "c2.cp.kegg.v7.2.symbols.gmt"
    gene_in_pathways = []
    for line in open(pathway):
        line = line.strip()
        split_result = line.split('\t')
        gene_in_pathway = split_result[2:]
        gene_in_pathways.append(gene_in_pathway)
    return gene_in_pathways


def get_col_value(wsheet, col):
    rows = wsheet.max_row
    col_data = []
    for i in range(2, rows + 1):
        cell_value = wsheet.cell(i, col).value
        col_data.append(cell_value)
    return col_data


def comb_mine(n, k):
    up = np.math.factorial(n)
    down = np.math.factorial(k) * np.math.factorial((n - k))
    return up // down


label6269_1 = list(chain.from_iterable(pd.read_excel('label6269-1.xlsx', usecols=[2]).values))
print(label6269_1)
print(len(label6269_1))
label6269_2 = list(chain.from_iterable(pd.read_excel('label6269-2.xlsx', usecols=[2]).values))
print(len(label6269_2))
label6269_3 = list(chain.from_iterable(pd.read_excel('label6269-3.xlsx', usecols=[2]).values))
print(len(label6269_3))
label11755 = list(chain.from_iterable(pd.read_excel('label11755.xlsx', usecols=[2]).values))
print(len(label11755))
label13015 = list(chain.from_iterable(pd.read_excel('label13015.xlsx', usecols=[2]).values))
label13015 = list(filter(lambda x: x != 10, label13015))
print(len(label13015))
label13015_2 = list(chain.from_iterable(pd.read_excel('label13015-2.xlsx', usecols=[2]).values))
label13015_2 = list(filter(lambda x: x != 10, label13015_2))
print(len(label13015_2))
label16129 = list(chain.from_iterable(pd.read_excel('label16129.xlsx', usecols=[2]).values))
print(len(label16129))
label16129_2 = list(chain.from_iterable(pd.read_excel('label16129-2.xlsx', usecols=[2]).values))
print(len(label16129_2))
label16129_3 = list(chain.from_iterable(pd.read_excel('label16129-3.xlsx', usecols=[2]).values))
print(len(label16129_3))
label17156 = list(chain.from_iterable(pd.read_excel('label17156.xlsx', usecols=[2]).values))
print(len(label17156))
label22098 = list(chain.from_iterable(pd.read_excel('label22098.xlsx', usecols=[2]).values))
label22098 = list(filter(lambda x: x != 10, label22098))
print(len(label22098))
label23140 = list(chain.from_iterable(pd.read_excel('label23140.xlsx', usecols=[2]).values))
print(len(label23140))
label25504 = list(chain.from_iterable(pd.read_excel('label25504.xlsx', usecols=[2]).values))
label25504 = list(filter(lambda x: x != 14, label25504))
print(len(label25504))
label25504_3 = list(chain.from_iterable(pd.read_excel('label25504-3.xlsx', usecols=[2]).values))
label25504_3 = list(filter(lambda x: x != 3, label25504_3))
print(len(label25504_3))
label25504_4 = list(chain.from_iterable(pd.read_excel('label25504-4.xlsx', usecols=[2]).values))
label25504_4 = list(filter(lambda x: x != 14, label25504_4))
print(len(label25504_4))
label29161 = list(chain.from_iterable(pd.read_excel('label29161.xlsx', usecols=[2]).values))
print(len(label29161))
label33341_2 = list(chain.from_iterable(pd.read_excel('label33341-2.xlsx', usecols=[2]).values))
print(len(label33341_2))
label34205 = list(chain.from_iterable(pd.read_excel('label34205.xlsx', usecols=[2]).values))
print(len(label34205))
label38246 = list(chain.from_iterable(pd.read_excel('label38246.xlsx', usecols=[2]).values))
print(len(label38246))
label38900 = list(chain.from_iterable(pd.read_excel('label38900.xlsx', usecols=[2]).values))
print(len(label38900))
label40586 = list(chain.from_iterable(pd.read_excel('label40586.xlsx', usecols=[2]).values))
print(len(label40586))
label51808 = list(chain.from_iterable(pd.read_excel('label51808.xlsx', usecols=[2]).values))
label51808 = list(filter(lambda x: x != 10, label51808))
print(len(label51808))
label69606 = list(chain.from_iterable(pd.read_excel('label69606.xlsx', usecols=[2]).values))
print(len(label69606))
label42834 = list(chain.from_iterable(pd.read_excel('label42834.xlsx', usecols=[2]).values))
label42834 = list(filter(lambda x: x != 10, label42834))
print(len(label42834))
label6377 = list(chain.from_iterable(pd.read_excel('label6377.xlsx', usecols=[2]).values))

excel40396 = openpyxl.load_workbook('label40396.xlsx')
sheets1 = excel40396.get_sheet_names()
ws1 = excel40396.get_sheet_by_name(sheets1[0])
label40396 = get_col_value(ws1, 3)

all_need_genes = load_pathway()
all_genes = pd.read_excel('all_data1.xlsx')
print(all_genes.index)
print(all_genes.columns)

label_excel = xlrd.open_workbook('all_labels.xls')
sheet0 = label_excel.sheets()[12]
label68310 = sheet0.col_values(2, 1, )
sheet1 = label_excel.sheets()[11]
label20346 = sheet1.col_values(2, 1, )
print(len(label20346))
sheet2 = label_excel.sheets()[10]
label21802 = sheet2.col_values(2, 1, )
print(len(label21802))
sheet3 = label_excel.sheets()[9]
label27131 = sheet3.col_values(2, 1, )
print(len(label27131))
sheet4 = label_excel.sheets()[8]
label28750 = sheet4.col_values(2, 1, )
print(len(label28750))
sheet5 = label_excel.sheets()[7]
label40012 = sheet5.col_values(2, 1, )
label40012 = list(filter(lambda x: x != 3, label40012))
print(len(label40012))
sheet6 = label_excel.sheets()[6]
label42026 = sheet6.col_values(2, 1, )
print(len(label42026))
sheet7 = label_excel.sheets()[5]
label57065 = sheet7.col_values(2, 1, )
print(len(label57065))
sheet8 = label_excel.sheets()[4]
label60244 = sheet8.col_values(2, 1, )
# label60244 = list(filter(lambda x: x != 3, label60244))
print(len(label60244))
sheet9 = label_excel.sheets()[3]
label63990 = sheet9.col_values(2, 1, )
print(len(label63990))
sheet10 = label_excel.sheets()[2]
label66099 = sheet10.col_values(2, 1, )
print(len(label66099))
sheet11 = label_excel.sheets()[1]
label69528 = sheet11.col_values(2, 1, )
print(len(label69528))
sheet12 = label_excel.sheets()[0]
label111368 = sheet12.col_values(2, 1, )
print(len(label111368))
label9692 = list(chain.from_iterable(pd.read_excel('label9692.xlsx', usecols=[2]).values))
label29385 = list(chain.from_iterable(pd.read_excel('label29385.xlsx', usecols=[2]).values))
label33341 = list(chain.from_iterable(pd.read_excel('label33341.xlsx', usecols=[2]).values))
label61821 = list(chain.from_iterable(pd.read_excel('label61821.xlsx', usecols=[2]).values))
label68004 = list(chain.from_iterable(pd.read_excel('label68004.xlsx', usecols=[2]).values))
label103842 = list(chain.from_iterable(pd.read_excel('label103842.xlsx', usecols=[2]).values))
labels = np.array(
    label6269_1 + label6269_2 + label6269_3 + label11755 + label13015_2 + label16129 + label16129_2 +
    label17156 + label20346 + label22098 + label23140 + label25504 + label25504_3 + label25504_4 + label27131 + label28750 + label29161 + label33341_2
    + label34205 + label38246 + label38900 + label40012 + label40396 + label42026 + label42834 + label51808 + label57065 +
    label60244 + label63990 + label66099 +
    label68310 + label69528 + label69606  + label111368)
print(len(labels))

label_health = []
label_num = 0
for i in all_genes.index:
    label_health.append(labels[i])
label_health = np.array(label_health)
label_h = []
for i in label_health:
    if i == 0 or i == 4:
        j = 1
    else:
        j = 0
    label_h.append(j)
label_h = np.array(label_h)

label_bac = []
bac_index = []
bac_num = 0
for i in labels:
    if i == 0 or i == 11 or i == 12 or i == 4 or i == 2:
        if bac_num in all_genes.index:
            bac_index.append(bac_num)
            j = i
            label_bac.append(j)
    bac_num = bac_num + 1
print(bac_index)
print(len(bac_index))
label_b = []
for i in label_bac:
    if i == 11:
        j = 1
    else:
        j = 0
    label_b.append(j)
label_b = np.array(label_b)
print(label_b)

label_bacn = []
bacn_index = []
bacn_num = 0
for i in labels:
    if i == 0 or i == 11 or i == 12 or i == 4 or i == 2:
        if bacn_num in all_genes.index:
            bacn_index.append(bacn_num)
            j = i
            label_bacn.append(j)
    bacn_num = bacn_num + 1
print(bacn_index)
print(len(bacn_index))
label_bn = []
for i in label_bacn:
    if i == 12:
        j = 1
    else:
        j = 0
    label_bn.append(j)
label_bn = np.array(label_bn)
print(label_bn)

label_virus = []
for i in all_genes.index:
    label_virus.append(labels[i])
label_virus = np.array(label_virus)
label_v = []
for i in label_virus:
    if i == 2:
        j = 1
    else:
        j = 0
    label_v.append(j)
label_v = np.array(label_v)

health_data = []
h_genes = []
for health_gene in all_health_genes:
    gene1 = health_gene[0]
    gene2 = health_gene[1]
    pair = []
    if gene1 in all_genes.columns and gene2 in all_genes.columns:
        each_pair = np.array(all_genes[gene1]) > np.array(all_genes[gene2])
        for i in each_pair:
            if i == 1:
                j = 1
            else:
                j = -1
            pair.append(j)
        health_data.append(pair)
        h_genes.append(health_gene)

health_data = np.array(health_data)
print(health_data.shape)
print(health_data)
health_data = health_data.T
print(health_data.shape)

bac_data = []
p_genes = []
for bac_gene in all_positive_genes:
    gene1 = bac_gene[0]
    gene2 = bac_gene[1]
    pair = []
    if gene1 in all_genes.columns and gene2 in all_genes.columns:
        each_pair = np.array(all_genes.loc[bac_index, gene1].values) > np.array(all_genes.loc[bac_index, gene2].values)
        each_pair = each_pair.flatten()
        for i in each_pair:
            if i == 1:
                j = 1
            else:
                j = -1
            pair.append(j)
        bac_data.append(pair)
        h_genes.append(bac_gene)
bac_data = np.array(bac_data)
bac_data = bac_data.T

bacn_data = []
n_genes = []
for bacn_gene in all_negative_genes:
    gene1 = bacn_gene[0]
    gene2 = bacn_gene[1]
    pair = []
    if gene1 in all_genes.columns and gene2 in all_genes.columns:
        each_pair = np.array(all_genes.loc[bacn_index, gene1].values) > np.array(
            all_genes.loc[bacn_index, gene2].values)
        each_pair = each_pair.flatten()
        for i in each_pair:
            if i == 1:
                j = 1
            else:
                j = -1
            pair.append(j)
        bacn_data.append(pair)
        h_genes.append(bacn_gene)
bacn_data = np.array(bacn_data)
bacn_data = bacn_data.T

virus_data = []
v_genes = []
for virus_gene in all_virus_genes:
    gene1 = virus_gene[0]
    gene2 = virus_gene[1]
    pair = []
    if gene1 in all_genes.columns and gene2 in all_genes.columns:
        each_pair = np.array(all_genes[gene1]) > np.array(
            all_genes[gene2])
        each_pair = each_pair.flatten()
        for i in each_pair:
            if i == 1:
                j = 1
            else:
                j = -1
            pair.append(j)
        virus_data.append(pair)
        h_genes.append(virus_gene)


virus_data = np.array(virus_data)
virus_data = virus_data.T



lasso1 = Lasso(alpha=0.05).fit(health_data, label_h)
health_result = (lasso1.coef_ * 100).tolist()
print(len(health_result))
lasso2 = Lasso(alpha=0.05).fit(bac_data, label_b)
bac_result = (lasso2.coef_ * 100).tolist()

lasso3 = Lasso(alpha=0.05).fit(bacn_data, label_bn)
bacn_result = (lasso3.coef_ * 100).tolist()

lasso4 = Lasso(alpha=0.05).fit(virus_data, label_v)
virus_result = (lasso4.coef_ * 100).tolist()

for i in health_result:
    with open('lasso_coe1.txt', 'a', encoding='utf-8') as h:
        h.write(str(abs(i)))
        h.write('\n')
for j in bac_result:
    with open('lasso_coe2', 'a', encoding='utf-8') as b:
        b.write(str(abs(j)))
        b.write('\n')
for z in virus_result:
    with open('lasso_coe3', 'a', encoding='utf-8') as v:
        v.write(str(abs(z)))
        v.write('\n')
for x in bacn_result:
    with open('lasso_coe4', 'a', encoding='utf-8') as bn:
        bn.write(str(abs(x)))
        bn.write('\n')

